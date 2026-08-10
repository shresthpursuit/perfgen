"""Synthetic specification workbooks for the parser tests.

The tool is never run against a real NFR spreadsheet while building (see CLAUDE.md), so the parser
is exercised against workbooks generated here. These mirror the structure of the shipped
`data/specs/*.xlsx`: the same sheet names, header text, field labels, section dividers and
footnote rows.

Two deliberate properties, both of which make a wrong parser fail rather than pass by luck:

* the `Example` column holds values that DIFFER from the `Value` column on every row, so a parser
  reading column E produces visibly wrong output;
* every knob a real user might turn - inserted rows, renamed sections, reordered columns, missing
  fields - is a parameter, so the mangled-spec test is the same builder with different arguments.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------------------------
# Default content - a valid, complete specification
# --------------------------------------------------------------------------------------------

APPLICATION_SECTION = "Application"
AUTHENTICATION_SECTION = "Authentication"

# (label, value, example) in sheet order. `None` value means the user left the cell empty.
DEFAULT_APPLICATION: list[tuple[str, str | int | None, str]] = [
    ("Application name", "Claims intake service", "Service request portal"),
    ("Base URL", "https://perf-claims.example.internal", "https://perf-api.example.internal"),
    ("Base path", "/v1", "/api/v2"),
    ("API reference location", None, "https://perf-api.example.internal/swagger.json"),
    ("Auth type", "OAuth2 client credentials", "OAuth2 client credentials"),
    ("Token endpoint URL", "https://sso.example.internal/connect/token", "https://auth.x/token"),
    ("Token request method", "POST", "POST"),
    (
        "Token request content type",
        "application/x-www-form-urlencoded",
        "application/x-www-form-urlencoded",
    ),
    ("Token request parameters", "grant_type\nclient_id\nclient_secret", "grant_type\nscope"),
    ("Token lifetime (seconds)", 3600, "3600"),
    ("Auth header name", "Authorization", "Authorization"),
    ("Auth header value format", "Bearer {token}", "Bearer {token}"),
    ("Credential reference names", "claims-perf-id\nclaims-perf-secret", "perf-client-id"),
    ("Account model", "Single shared", "Single shared"),
]

# Labels that sit under the Authentication divider rather than the Application one.
AUTH_LABELS = {label for label, _, _ in DEFAULT_APPLICATION[4:]}

APPLICATION_HEADERS = ["Attribute", "Value", "Required", "Notes", "Example"]
FLOW_HEADERS = [
    "Flow ID",
    "Flow name",
    "Share of load %",
    "Think time between calls (s)",
    "Safe to run against this environment",
]
STEP_HEADERS = [
    "Flow ID",
    "Step no",
    "Step name",
    "Method",
    "Endpoint path",
    "Request body or parameters",
    "Expected status",
]
PROFILE_HEADERS = [
    "Test type",
    "Required",
    "Concurrent users",
    "Ramp-up (s)",
    "Duration (min)",
    "Target throughput",
    "Throughput unit",
]
SLA_HEADERS = ["Applies to", "Metric", "Target", "Unit"]

DEFAULT_FLOWS = [
    ["F01", "Search and view record", 60, 3, "Yes"],
    ["F02", "Submit new request", 40, 5, "No"],
]

DEFAULT_STEPS = [
    ["F01", 1, "Search catalogue", "GET", "/catalogue/search?q=widget", None, 200],
    ["F01", 2, "Open record detail", "GET", "/catalogue/items/{itemId}", None, 200],
    ["F02", 1, "Create request", "POST", "/requests", '{"type":"standard"}', 201],
    ["F02", 2, "Check status", "GET", "/requests/{requestRef}/status", None, 200],
]

DEFAULT_PROFILES = [
    ["Baseline", "Yes", 25, 120, 30, 1200, "TPH"],
    ["Peak load", "Yes", 80, 180, 30, 4000, "TPH"],
    ["Capacity / overload", "No", None, None, None, None, None],
    ["Endurance", "Yes", 25, 120, 240, 1200, "TPH"],
]

DEFAULT_SLAS = [
    ["All flows", "Response time 95th percentile", 800, "ms"],
    ["F01", "Response time 99th percentile", 1500, "ms"],
    ["All flows", "Error rate", 1, "%"],
]

FOOTNOTE = (
    "Share of load must total 100 across all flows. Safe to run means the framework is allowed "
    "to call this flow once while building the script."
)


# --------------------------------------------------------------------------------------------
# Builder
# --------------------------------------------------------------------------------------------


@dataclass
class WorkbookSpec:
    """Everything a test might want to vary about a generated workbook."""

    application: list[tuple[str, str | int | None, str]] = field(
        default_factory=lambda: list(DEFAULT_APPLICATION)
    )
    flows: list[list] = field(default_factory=lambda: [list(r) for r in DEFAULT_FLOWS])
    steps: list[list] = field(default_factory=lambda: [list(r) for r in DEFAULT_STEPS])
    profiles: list[list] = field(default_factory=lambda: [list(r) for r in DEFAULT_PROFILES])
    slas: list[list] = field(default_factory=lambda: [list(r) for r in DEFAULT_SLAS])

    # Sheet naming
    sheet_names: dict[str, str] = field(
        default_factory=lambda: {
            "application": "Application",
            "flows": "Flows",
            "steps": "Flow steps",
            "profiles": "Load profiles",
            "sla": "SLA",
        }
    )
    include_read_me: bool = True

    # Mangling knobs
    blank_rows_before_header: int = 0
    """Rows inserted above the header row, shifting everything down."""

    extra_blank_rows_in_body: int = 0
    """Blank rows scattered between data rows."""

    section_titles: tuple[str, str] = (APPLICATION_SECTION, AUTHENTICATION_SECTION)
    """Section divider text - users rename these."""

    column_order: dict[str, list[str]] = field(default_factory=dict)
    """Per-sheet header order override, e.g. {"flows": ["Flow name", "Flow ID", ...]}."""

    omit_columns: dict[str, list[str]] = field(default_factory=dict)
    """Per-sheet headers to drop entirely."""

    include_example_column: bool = True
    include_footnotes: bool = True
    include_dropdowns: bool = True


def _headers_for(sheet: str, defaults: list[str], spec: WorkbookSpec) -> list[str]:
    headers = list(spec.column_order.get(sheet, defaults))
    for omitted in spec.omit_columns.get(sheet, []):
        if omitted in headers:
            headers.remove(omitted)
    return headers


def _write_table(ws, sheet_key: str, headers: list[str], rows: list[list], spec: WorkbookSpec):
    """Write a header row then data rows, honouring column reordering and blank-row insertion."""
    ordered = _headers_for(sheet_key, headers, spec)
    index_of = {name: headers.index(name) for name in headers}

    row_cursor = 1 + spec.blank_rows_before_header
    for column, name in enumerate(ordered, start=1):
        ws.cell(row=row_cursor, column=column, value=name)

    row_cursor += 1
    for row in rows:
        for column, name in enumerate(ordered, start=1):
            value = row[index_of[name]]
            if value is not None:
                ws.cell(row=row_cursor, column=column, value=value)
        row_cursor += 1 + spec.extra_blank_rows_in_body

    if spec.include_footnotes:
        ws.cell(row=row_cursor + 2, column=1, value=FOOTNOTE)
    return ordered, row_cursor


def _write_application(ws, spec: WorkbookSpec) -> None:
    headers = list(APPLICATION_HEADERS)
    if not spec.include_example_column:
        headers.remove("Example")
    ordered = _headers_for("application", headers, spec)

    row = 1 + spec.blank_rows_before_header
    for column, name in enumerate(ordered, start=1):
        ws.cell(row=row, column=column, value=name)
    row += 1

    app_title, auth_title = spec.section_titles
    attribute_col = ordered.index("Attribute") + 1 if "Attribute" in ordered else 1

    ws.cell(row=row, column=attribute_col, value=app_title)
    row += 1

    seen_auth_divider = False
    for label, value, example in spec.application:
        if label in AUTH_LABELS and not seen_auth_divider:
            ws.cell(row=row, column=attribute_col, value=auth_title)
            row += 1
            seen_auth_divider = True

        cells = {
            "Attribute": label,
            "Value": value,
            "Required": "Yes",
            "Notes": "Guidance text for the user.",
            "Example": example,
        }
        for column, name in enumerate(ordered, start=1):
            cell_value = cells.get(name)
            if cell_value is not None:
                ws.cell(row=row, column=column, value=cell_value)
        row += 1 + spec.extra_blank_rows_in_body

    if spec.include_footnotes:
        ws.cell(
            row=row + 1,
            column=1,
            value="Never type a password, secret or key into this workbook.",
        )

    if spec.include_dropdowns and "Value" in ordered:
        value_letter = chr(64 + ordered.index("Value") + 1)
        _add_validation(
            ws,
            f"{value_letter}2:{value_letter}40",
            '"None,OAuth2 client credentials,OAuth2 password,OAuth2 PKCE,'
            'Bearer static,API key,Basic"',
        )


def _add_validation(ws, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def write_workbook(path: str | Path, spec: WorkbookSpec | None = None) -> Path:
    """Generate a specification workbook at `path`."""
    spec = spec or WorkbookSpec()
    workbook = Workbook()
    workbook.remove(workbook.active)

    if spec.include_read_me:
        read_me = workbook.create_sheet("Read me")
        read_me.cell(row=2, column=2, value="API performance test specification")
        read_me.cell(row=3, column=2, value="Fill in the five sheets that follow.")

    _write_application(workbook.create_sheet(spec.sheet_names["application"]), spec)

    flows = workbook.create_sheet(spec.sheet_names["flows"])
    _write_table(flows, "flows", FLOW_HEADERS, spec.flows, spec)

    steps = workbook.create_sheet(spec.sheet_names["steps"])
    _write_table(steps, "steps", STEP_HEADERS, spec.steps, spec)

    profiles = workbook.create_sheet(spec.sheet_names["profiles"])
    _write_table(profiles, "profiles", PROFILE_HEADERS, spec.profiles, spec)

    sla = workbook.create_sheet(spec.sheet_names["sla"])
    _write_table(sla, "sla", SLA_HEADERS, spec.slas, spec)

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


# --------------------------------------------------------------------------------------------
# Convenience mutators for tests
# --------------------------------------------------------------------------------------------


def set_application_value(spec: WorkbookSpec, label: str, value: str | int | None) -> WorkbookSpec:
    """Overwrite one Application field's value, leaving its Example cell untouched."""
    spec.application = [
        (existing_label, value if existing_label == label else existing_value, example)
        for existing_label, existing_value, example in spec.application
    ]
    return spec


def drop_application_row(spec: WorkbookSpec, label: str) -> WorkbookSpec:
    """Remove a label from the sheet entirely, as if the user deleted the row."""
    spec.application = [row for row in spec.application if row[0] != label]
    return spec


def profile_row(spec: WorkbookSpec, test_type: str) -> list:
    return next(row for row in spec.profiles if row[0] == test_type)
