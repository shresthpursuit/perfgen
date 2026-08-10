"""Locating things in a worksheet by what they say, never by where they sit.

Users insert rows, rename tabs, reorder columns and delete the ones they did not need.
Coordinate-based parsing breaks silently on the second file you receive, and silently is the
problem: it reads a Notes cell as a Base URL and generates a script that runs against nothing.

Everything here matches on normalised text and reports what it could not find.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

_WHITESPACE = re.compile(r"\s+")


def normalise(text: Any) -> str:
    """Casefold and collapse whitespace, so 'Base  URL ' matches 'base url'."""
    if text is None:
        return ""
    return _WHITESPACE.sub(" ", str(text).strip()).casefold()


def cell_text(value: Any) -> str | None:
    """Trim a cell to a string, treating blank and whitespace-only as absent."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def find_sheet(workbook, *candidates: str) -> Worksheet | None:
    """Find a sheet by name, tolerating case and spacing differences."""
    wanted = {normalise(name) for name in candidates}
    for name in workbook.sheetnames:
        if normalise(name) in wanted:
            return workbook[name]
    return None


@dataclass
class HeaderRow:
    """A located header row and the column index of each header it carries."""

    row_index: int
    columns: dict[str, int]  # normalised header text -> 1-based column index

    def column_of(self, *names: str) -> int | None:
        for name in names:
            index = self.columns.get(normalise(name))
            if index is not None:
                return index
        return None

    def has(self, *names: str) -> bool:
        return self.column_of(*names) is not None


def find_header_row(
    sheet: Worksheet, expected: list[str], *, search_limit: int = 25
) -> HeaderRow | None:
    """Find the row that carries the table headers.

    The header is row 1 in the shipped template, but a user who inserts rows above it should not
    break parsing, so the first `search_limit` rows are scanned for the best match.
    """
    wanted = {normalise(name) for name in expected}
    best: HeaderRow | None = None
    best_score = 0

    for row in sheet.iter_rows(min_row=1, max_row=search_limit):
        columns: dict[str, int] = {}
        for cell in row:
            text = normalise(cell.value)
            if text and text not in columns:
                columns[text] = cell.column
        score = len(wanted & set(columns))
        if score > best_score:
            best_score, best = score, HeaderRow(row[0].row, columns)

    # One matching header could be a coincidence in a prose row; two is a table.
    if best is None or best_score < min(2, len(wanted)):
        return None
    return best


def iter_data_rows(sheet: Worksheet, header: HeaderRow, *, key_column: int):
    """Yield (row_index, row) for rows below the header that carry a value in `key_column`.

    Blank spacer rows are skipped rather than ending the table, and the trailing footnote rows the
    template carries are skipped too: a footnote is a long sentence in the first column with every
    other column empty.
    """
    for row in sheet.iter_rows(min_row=header.row_index + 1):
        cells = {cell.column: cell.value for cell in row}
        key = cell_text(cells.get(key_column))
        if key is None:
            continue
        populated = [c for c, v in cells.items() if cell_text(v) is not None]
        if populated == [key_column] and _looks_like_prose(key):
            continue
        yield row[0].row, cells


def _looks_like_prose(text: str) -> bool:
    """A footnote is a sentence; a Flow ID or a test type is not."""
    return len(text) > 60 and " " in text


@dataclass
class LabelledValue:
    """One attribute/value pair located on a key-value sheet."""

    row_index: int
    value: Any


class LabelledSheet:
    """An attribute/value sheet, e.g. Application.

    Values are read from the column whose header says `Value` - deliberately not "the cell next to
    the label", because the sheet also carries an `Example` column whose contents differ from the
    real answer. Reading the wrong column yields a plausible-looking wrong spec.
    """

    def __init__(self, sheet: Worksheet, header: HeaderRow):
        self.sheet = sheet
        self.header = header
        self.attribute_column = header.column_of("Attribute", "Field", "Setting") or 1
        self.value_column = header.column_of("Value")
        self._index: dict[str, LabelledValue] = {}

        if self.value_column is None:
            return

        for row in sheet.iter_rows(min_row=header.row_index + 1):
            cells = {cell.column: cell.value for cell in row}
            label = cell_text(cells.get(self.attribute_column))
            if label is None:
                continue
            key = normalise(label)
            if key not in self._index:
                self._index[key] = LabelledValue(row[0].row, cells.get(self.value_column))

    @property
    def usable(self) -> bool:
        return self.value_column is not None

    def get(self, label: str) -> Any:
        entry = self._index.get(normalise(label))
        return entry.value if entry else None

    def has_label(self, label: str) -> bool:
        return normalise(label) in self._index

    def row_of(self, label: str) -> int | None:
        entry = self._index.get(normalise(label))
        return entry.row_index if entry else None
